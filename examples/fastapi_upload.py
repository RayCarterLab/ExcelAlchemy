"""Minimal FastAPI integration example for template download and workbook import."""

import asyncio
import io
from base64 import b64decode
from io import BytesIO

from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from excelalchemy import (
    ExcelAlchemy,
    ExcelStorage,
    FieldMeta,
    ImporterConfig,
    ImportResult,
    Number,
    String,
    UrlStr,
)
from excelalchemy.core.table import WorksheetTable


class EmployeeImporter(BaseModel):
    full_name: String = FieldMeta(label='Full name', order=1, hint='Use the legal name')
    age: Number = FieldMeta(label='Age', order=2)


class RequestScopedStorage(ExcelStorage):
    """Self-contained request-scoped storage used by the FastAPI example."""

    def __init__(self) -> None:
        self.fixtures: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}

    def register_upload(self, input_excel_name: str, content: bytes) -> None:
        self.fixtures[input_excel_name] = content

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')


async def create_employee(row: dict[str, object], context: dict[str, object] | None) -> dict[str, object]:
    if context is not None:
        created_rows = context.setdefault('created_rows', [])
        assert isinstance(created_rows, list)
        created_rows.append(row.copy())
        row['tenant_id'] = context['tenant_id']
    return row


def build_import_alchemy(
    storage: RequestScopedStorage, *, tenant_id: str = 'tenant-001'
) -> ExcelAlchemy[dict[str, object], EmployeeImporter, EmployeeImporter]:
    alchemy = ExcelAlchemy(
        ImporterConfig.for_create(
            EmployeeImporter,
            creator=create_employee,
            storage=storage,
            locale='en',
        )
    )
    alchemy.add_context({'tenant_id': tenant_id, 'created_rows': []})
    return alchemy


def build_template_response() -> StreamingResponse:
    alchemy = ExcelAlchemy(ImporterConfig.for_create(EmployeeImporter, locale='en'))
    artifact = alchemy.download_template_artifact(filename='employee-template.xlsx')
    return StreamingResponse(
        BytesIO(artifact.as_bytes()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=employee-template.xlsx'},
    )


async def import_employees_from_upload(storage: RequestScopedStorage, file: UploadFile) -> dict[str, object]:
    if not file.filename:
        raise HTTPException(status_code=400, detail='An Excel file is required')

    storage.register_upload(file.filename, await file.read())
    alchemy = build_import_alchemy(storage)
    result = await alchemy.import_data(file.filename, 'employee-import-result.xlsx')
    created_rows = alchemy.context['created_rows']
    assert isinstance(created_rows, list)
    return {
        'result': result.model_dump(),
        'created_rows': len(created_rows),
        'uploaded_artifacts': sorted(storage.uploaded),
    }


def create_app(storage: RequestScopedStorage | None = None) -> FastAPI:
    app = FastAPI()
    request_storage = storage or RequestScopedStorage()

    @app.get('/employee-template.xlsx')
    async def download_template() -> StreamingResponse:
        return build_template_response()

    @app.post('/employee-imports')
    async def import_employees(file: UploadFile) -> dict[str, object]:
        return await import_employees_from_upload(request_storage, file)

    return app


def _build_demo_upload(template_bytes: bytes) -> bytes:
    workbook = load_workbook(io.BytesIO(template_bytes))
    try:
        worksheet = workbook['Sheet1']
        worksheet['A3'] = 'TaylorChen'
        worksheet['B3'] = '32'
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    finally:
        workbook.close()


async def run_demo() -> tuple[ImportResult, RequestScopedStorage, int]:
    template_alchemy = ExcelAlchemy(ImporterConfig.for_create(EmployeeImporter, locale='en'))
    template = template_alchemy.download_template_artifact(filename='employee-template.xlsx')
    storage = RequestScopedStorage()
    upload_bytes = _build_demo_upload(template.as_bytes())
    upload_file = UploadFile(filename='employee-import.xlsx', file=BytesIO(upload_bytes))
    response_payload = await import_employees_from_upload(storage, upload_file)
    result = ImportResult.model_validate(response_payload['result'])
    created_rows = response_payload['created_rows']
    assert isinstance(created_rows, int)
    return result, storage, created_rows


def main() -> None:
    result, storage, created_rows = asyncio.run(run_demo())
    app = create_app(storage)
    route_paths = sorted(route.path for route in app.routes if getattr(route, 'path', None))

    print('FastAPI upload example completed')
    print(f'Result: {result.result}')
    print(f'Success rows: {result.success_count}')
    print(f'Failed rows: {result.fail_count}')
    print(f'Created rows: {created_rows}')
    print(f'Uploaded artifacts: {sorted(storage.uploaded)}')
    print(f'Routes: {route_paths}')


app = create_app()


__all__ = [
    'EmployeeImporter',
    'RequestScopedStorage',
    'app',
    'build_import_alchemy',
    'build_template_response',
    'create_app',
    'import_employees_from_upload',
    'main',
    'run_demo',
]


if __name__ == '__main__':
    main()
