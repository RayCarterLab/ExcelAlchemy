"""Request-scoped storage used by the FastAPI reference project."""

import io
from base64 import b64decode

from openpyxl import load_workbook

from excelalchemy import ExcelStorage, UrlStr
from excelalchemy.core.table import WorksheetTable


class RequestScopedStorage(ExcelStorage):
    """Self-contained in-memory storage for uploads and result workbooks."""

    def __init__(self) -> None:
        self.fixtures: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}

    def register_upload(self, input_excel_name: str, content: bytes) -> None:
        self.fixtures[input_excel_name] = content

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')


__all__ = ['RequestScopedStorage']
