"""Create-or-update import example with explicit existence checks."""

import asyncio
import io
from base64 import b64decode

from openpyxl import load_workbook
from pydantic import BaseModel

from excelalchemy import Email, ExcelAlchemy, ExcelStorage, FieldMeta, ImporterConfig, ImportResult, String, UrlStr
from excelalchemy.core.table import WorksheetTable


class CustomerImporter(BaseModel):
    customer_name: String = FieldMeta(label='Customer name', order=1)
    work_email: Email = FieldMeta(label='Work email', order=2)
    team: String = FieldMeta(label='Team', order=3)


class InMemoryUpsertStorage(ExcelStorage):
    def __init__(self) -> None:
        self.fixtures: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')


def _build_import_fixture(storage: InMemoryUpsertStorage, template_bytes: bytes) -> None:
    workbook = load_workbook(io.BytesIO(template_bytes))
    try:
        worksheet = workbook['Sheet1']
        worksheet['A3'] = 'TaylorChen'
        worksheet['B3'] = 'taylor.chen@example.com'
        worksheet['C3'] = 'Finance'

        worksheet['A4'] = 'AveryStone'
        worksheet['B4'] = 'avery.stone@example.com'
        worksheet['C4'] = 'Operations'

        buffer = io.BytesIO()
        workbook.save(buffer)
        storage.fixtures['customer-import.xlsx'] = buffer.getvalue()
    finally:
        workbook.close()


async def is_customer_existing(row: dict[str, object], context: dict[str, object] | None) -> bool:
    if context is None:
        return False
    existing_emails = context.setdefault('existing_emails', set())
    assert isinstance(existing_emails, set)
    email = str(row['work_email'])
    return email in existing_emails


async def create_customer(row: dict[str, object], context: dict[str, object] | None) -> dict[str, object]:
    if context is not None:
        created_rows = context.setdefault('created_rows', [])
        existing_emails = context.setdefault('existing_emails', set())
        assert isinstance(created_rows, list)
        assert isinstance(existing_emails, set)
        created_rows.append(row.copy())
        existing_emails.add(str(row['work_email']))
    return row


async def update_customer(row: dict[str, object], context: dict[str, object] | None) -> dict[str, object]:
    if context is not None:
        updated_rows = context.setdefault('updated_rows', [])
        assert isinstance(updated_rows, list)
        updated_rows.append(row.copy())
    return row


async def run_workflow() -> tuple[ImportResult, InMemoryUpsertStorage, dict[str, object]]:
    storage = InMemoryUpsertStorage()
    context: dict[str, object] = {
        'existing_emails': {'taylor.chen@example.com'},
        'created_rows': [],
        'updated_rows': [],
    }

    alchemy = ExcelAlchemy(
        ImporterConfig.for_create_or_update(
            create_importer_model=CustomerImporter,
            update_importer_model=CustomerImporter,
            is_data_exist=is_customer_existing,
            creator=create_customer,
            updater=update_customer,
            storage=storage,
            locale='en',
        )
    )
    alchemy.add_context(context)

    template = alchemy.download_template_artifact(filename='customer-template.xlsx')
    _build_import_fixture(storage, template.as_bytes())
    result = await alchemy.import_data('customer-import.xlsx', 'customer-import-result.xlsx')
    return result, storage, context


def main() -> None:
    result, storage, context = asyncio.run(run_workflow())
    created_rows = context['created_rows']
    updated_rows = context['updated_rows']
    assert isinstance(created_rows, list)
    assert isinstance(updated_rows, list)

    print('Create-or-update import workflow completed')
    print(f'Result: {result.result}')
    print(f'Success rows: {result.success_count}')
    print(f'Failed rows: {result.fail_count}')
    print(f'Created rows: {len(created_rows)}')
    print(f'Updated rows: {len(updated_rows)}')
    print(f'Result workbook URL: {result.url}')
    print(f'Uploaded artifacts: {sorted(storage.uploaded)}')


if __name__ == '__main__':
    main()
