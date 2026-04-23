"""End-to-end import example with template generation and in-memory workbook uploads."""

import asyncio
import io
from base64 import b64decode

from openpyxl import load_workbook
from pydantic import BaseModel

from excelalchemy import (
    Email,
    ExcelAlchemy,
    ExcelStorage,
    FieldMeta,
    ImporterConfig,
    ImportPreflightResult,
    ImportResult,
    Number,
    String,
    UrlStr,
)
from excelalchemy.core.table import WorksheetTable


class EmployeeImporter(BaseModel):
    full_name: String = FieldMeta(label='Full name', order=1, hint='Use the legal name')
    age: Number = FieldMeta(label='Age', order=2)
    work_email: Email = FieldMeta(label='Work email', order=3, hint='Use the company mailbox')


class InMemoryImportStorage(ExcelStorage):
    def __init__(self) -> None:
        self.fixtures: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')


def _build_import_fixture(storage: InMemoryImportStorage, template_bytes: bytes) -> None:
    workbook = load_workbook(io.BytesIO(template_bytes))
    try:
        worksheet = workbook['Sheet1']
        worksheet['A3'] = 'TaylorChen'
        worksheet['B3'] = '32'
        worksheet['C3'] = 'taylor.chen@example.com'

        buffer = io.BytesIO()
        workbook.save(buffer)
        storage.fixtures['employee-import.xlsx'] = buffer.getvalue()
    finally:
        workbook.close()


async def create_employee(row: dict[str, object], context: dict[str, object] | None) -> dict[str, object]:
    if context is not None:
        created_rows = context.setdefault('created_rows', [])
        assert isinstance(created_rows, list)
        created_rows.append(row.copy())
    return row


async def run_workflow() -> tuple[
    ImportPreflightResult,
    ImportResult,
    InMemoryImportStorage,
    dict[str, object],
    list[dict[str, object]],
]:
    storage = InMemoryImportStorage()
    context: dict[str, object] = {
        'created_rows': [],
        'job_progress': {
            'status': 'pending',
            'processed_rows': 0,
            'total_rows': 0,
        },
    }
    events: list[dict[str, object]] = []

    alchemy = ExcelAlchemy(
        ImporterConfig.for_create(
            EmployeeImporter,
            creator=create_employee,
            storage=storage,
            locale='en',
        )
    )
    alchemy.add_context(context)

    template = alchemy.download_template_artifact(filename='employee-template.xlsx')
    _build_import_fixture(storage, template.as_bytes())
    preflight = alchemy.preflight_import('employee-import.xlsx')
    assert preflight.is_valid

    def handle_import_event(event: dict[str, object]) -> None:
        events.append(event)
        job_progress = context['job_progress']
        assert isinstance(job_progress, dict)

        match event['event']:
            case 'started':
                job_progress['status'] = 'running'
            case 'row_processed':
                job_progress['processed_rows'] = event['processed_row_count']
                job_progress['total_rows'] = event['total_row_count']
            case 'completed':
                job_progress['status'] = 'completed'
                job_progress['result'] = event['result']
                job_progress['result_workbook_url'] = event['url']
            case 'failed':
                job_progress['status'] = 'failed'
                job_progress['error'] = event['error_message']

    result = await alchemy.import_data(
        'employee-import.xlsx',
        'employee-import-result.xlsx',
        on_event=handle_import_event,
    )
    return preflight, result, storage, context, events


def main() -> None:
    preflight, result, storage, context, events = asyncio.run(run_workflow())
    created_rows = context['created_rows']
    job_progress = context['job_progress']
    assert isinstance(created_rows, list)
    assert isinstance(job_progress, dict)

    print('Employee import workflow completed')
    print(f'Preflight: {preflight.status}')
    print(f'Result: {result.result}')
    print(f'Success rows: {result.success_count}')
    print(f'Failed rows: {result.fail_count}')
    print(f'Result workbook URL: {result.url}')
    print(f'Created rows: {len(created_rows)}')
    print(f'Uploaded artifacts: {sorted(storage.uploaded)}')
    print(f'Observed events: {[event["event"] for event in events]}')
    print(f'Job progress: {job_progress}')


if __name__ == '__main__':
    main()
