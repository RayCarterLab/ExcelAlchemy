import base64
import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def decode_prefixed_excel_to_workbook(content: str) -> Workbook:
    _, payload = content.split(',', 1)
    return load_workbook(io.BytesIO(base64.b64decode(payload)))


def load_binary_excel_to_workbook(content: bytes) -> Workbook:
    return load_workbook(io.BytesIO(content))


def worksheet_matrix(
    worksheet: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> list[list[Any]]:
    return [
        [worksheet.cell(row=row_index, column=column_index).value for column_index in range(min_col, max_col + 1)]
        for row_index in range(min_row, max_row + 1)
    ]


def list_merge_ranges(worksheet: Worksheet) -> list[str]:
    return sorted(str(cell_range) for cell_range in worksheet.merged_cells.ranges)


def list_data_validations(worksheet: Worksheet) -> list[tuple[str | None, str]]:
    return [(validation.formula1, str(validation.sqref)) for validation in worksheet.data_validations.dataValidation]


def get_fill_color(cell: Cell | MergedCell) -> str | None:
    color = cell.fill.start_color.rgb or cell.fill.fgColor.rgb or cell.fill.start_color.index
    return _normalize_color(color)


def get_font_color(cell: Cell) -> str | None:
    if cell.font.color is None:
        return None
    color = cell.font.color.rgb or cell.font.color.indexed or cell.font.color.theme
    return _normalize_color(color)


def _normalize_color(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, str):
        upper = value.upper()
        return upper[-6:] if len(upper) == 8 else upper
    return str(value)
