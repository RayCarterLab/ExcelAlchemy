import io
import os
from copy import copy
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, ClassVar

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.cell_range import CellRange

from excelalchemy.const import HEADER_HINT
from tests.support.registry import FileRegistry


class LocalMockMinio:
    """有合并表头的内容直接使用文件"""

    storage: ClassVar[dict[str, Any]] = {}
    bucket_name: str = 'test'

    mock_excel_data: ClassVar[dict[str, Any]] = {
        FileRegistry.TEST_HEADER_INVALID_INPUT: [
            {
                '不存在的表头': '是',
            },
        ],
        FileRegistry.TEST_BOOLEAN_INPUT: [
            {
                '是否启用': '是',
            },
        ],
        FileRegistry.TEST_DATE_INPUT: [
            {
                '出生日期': '2021-01-01',
            },
        ],
        FileRegistry.TEST_DATE_INPUT_WRONG_RANGE: [
            {
                '出生日期': '2021-01-32',
            },
        ],
        FileRegistry.TEST_DATE_INPUT_WRONG_FORMAT: [
            {
                '出生日期': '2021-13',
            },
        ],
        FileRegistry.TEST_DATE_RANGE_INPUT: './files/test_date_range_input.xlsx',
        FileRegistry.TEST_DATE_RANGE_MISSING_INPUT_BEFORE: './files/test_date_range_missing_input_before.xlsx',
        FileRegistry.TEST_DATE_RANGE_MISSING_INPUT_AFTER: './files/test_date_range_missing_input_after.xlsx',
        FileRegistry.TEST_EMAIL_WRONG_FORMAT: [
            {
                '邮箱': '123',
            },
        ],
        FileRegistry.TEST_EMAIL_CORRECT_FORMAT: [
            {
                '邮箱': 'person@example.com',
            },
        ],
        FileRegistry.TEST_SIMPLE_IMPORT: [
            {
                '年龄': 18,
                '姓名': '张三',
                '地址': '北京市',
                '是否启用': '是',
                '出生日期': '2021-01-01',
                '邮箱': 'noreply@example.com',
                '价格': 100.0,
                '爱好': '篮球',
                '公司': '阿里巴巴',
                '经理': '李四',
                '部门': '研发部',
                '电话': '13800138000',
                '单选': '选项1',
                '老板': '马云',
                '领导': '张三',
                '团队': '研发部',
                '网址': 'https://www.baidu.com',
            }
        ],
        FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR: [
            {
                '年龄': '18',
                '姓名': 'Denial',
                '地址': '北京市',
                '是否启用': '是',
                '出生日期': '2021-01-32',
                '邮箱': 'not an email',
                '价格': '100.0',
                '爱好': '游泳',
                '公司': '亚马逊',
                '经理': '经理',
                '部门': '研发部门',
                '电话': '123',
                '单选': '选项10',
                '老板': '自己',
                '领导': '自己',
                '团队': '研发部门',
                '网址': 'google',
            }
        ],
        FileRegistry.TEST_IMPORT_WITH_MERGE_HEADER: './files/test_import_with_merge_header.xlsx',
    }

    def __init__(self):
        """generate mock (NamedTemporaryFile) Excel files from mock_excel_data

        automatically add HEADER_HINT to first row
        """
        for filename, data in self.mock_excel_data.items():
            with NamedTemporaryFile(suffix='.xlsx', delete=False) as temporary_file:
                temporary_filename = temporary_file.name

            workbook = self._build_workbook(data)
            workbook.save(temporary_filename)
            workbook.close()

            with open(temporary_filename, 'rb') as rf:
                file_bytes = rf.read()

            data = io.BytesIO(file_bytes)
            length = len(file_bytes)
            self.put_object(self.bucket_name, filename, data, length, temporary_filename)

    def _build_workbook(self, data: str | list[dict[str, Any]]):
        if isinstance(data, str):
            source_workbook = load_workbook(Path(__file__).resolve().parent.parent / Path(data.lstrip('./')))
            source_worksheet = source_workbook['Sheet1']
            rows = [list(row) for row in source_worksheet.iter_rows(values_only=True)]
            trimmed_width = self._trimmed_width(rows)

            workbook = Workbook()
            worksheet = workbook.active
            assert worksheet is not None
            worksheet.title = source_worksheet.title
            worksheet.cell(row=1, column=1, value=HEADER_HINT)

            for row_index, row in enumerate(rows, start=2):
                for column_index, value in enumerate(row[:trimmed_width], start=1):
                    worksheet.cell(row=row_index, column=column_index, value=value)

            for merged_range in source_worksheet.merged_cells.ranges:
                if merged_range.min_col > trimmed_width:
                    continue
                shifted_range = CellRange(
                    min_col=merged_range.min_col,
                    max_col=min(merged_range.max_col, trimmed_width),
                    min_row=merged_range.min_row + 1,
                    max_row=merged_range.max_row + 1,
                )
                worksheet.merge_cells(str(shifted_range))

            source_workbook.close()
            return workbook

        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = 'Sheet1'
        worksheet.cell(row=1, column=1, value=HEADER_HINT)

        if not data:
            return workbook

        headers = list(data[0].keys())
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=2, column=column_index, value=header)

        for row_index, row in enumerate(data, start=3):
            for column_index, header in enumerate(headers, start=1):
                worksheet.cell(row=row_index, column=column_index, value=row.get(header))

        return workbook

    @staticmethod
    def _trimmed_width(rows: list[list[Any]]) -> int:
        if not rows:
            return 0

        width = max(len(row) for row in rows)
        while width > 0:
            if any(len(row) >= width and row[width - 1] is not None for row in rows):
                return width
            width -= 1
        return 0

    def put_object(self, bucket_name: str, filename: str, data: io.BytesIO, length: int, file: Any = None) -> None:
        self.storage[filename] = {
            'bucket_name': bucket_name,
            'filename': filename,
            'data': data,
            'length': length,
            'file': file,
        }

    @classmethod
    def presigned_get_object(cls, bucket_name: str, filename: str, expires: int) -> str:
        return f'{bucket_name}/{filename}'

    def get_object(self, bucket_name: str, filename: str) -> io.BytesIO:
        assert bucket_name is not None
        return copy(self.storage[filename]['data'])  # use copy to avoid close(), so it can be read multiple times

    def __del__(self):
        for data in self.storage.values():
            if isinstance(data['file'], str) and os.path.exists(data['file']):
                os.remove(data['file'])


local_minio = LocalMockMinio()
