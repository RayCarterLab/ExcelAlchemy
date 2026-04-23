import io
from base64 import b64decode

from openpyxl import load_workbook

from excelalchemy import UrlStr
from excelalchemy.core.storage_protocol import ExcelStorage
from excelalchemy.core.table import WorksheetTable
from excelalchemy.exceptions import WorksheetNotFoundError
from excelalchemy.i18n.messages import MessageKey
from excelalchemy.i18n.messages import message as msg


class InMemoryExcelStorage(ExcelStorage):
    """Simple in-memory storage used to exercise the storage protocol."""

    def __init__(self, fixtures: dict[str, bytes] | None = None):
        self.fixtures = fixtures or {}
        self.uploaded: dict[str, bytes] = {}

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise WorksheetNotFoundError(
                    msg(MessageKey.WORKSHEET_NOT_FOUND, sheet_name=sheet_name),
                    message_key=MessageKey.WORKSHEET_NOT_FOUND,
                    sheet_name=sheet_name,
                )
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')
