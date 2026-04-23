import io
from typing import cast

from minio import Minio
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from excelalchemy import (
    ExcelAlchemy,
    FieldMeta,
    ImporterConfig,
    ImportPreflightStatus,
    String,
    ValidateResult,
    WorksheetNotFoundError,
)
from excelalchemy.const import BACKGROUND_ERROR_COLOR, REASON_COLUMN_LABEL, RESULT_COLUMN_LABEL
from excelalchemy.core.import_session import ImportSessionPhase
from excelalchemy.i18n.messages import MessageKey
from excelalchemy.i18n.messages import message as msg
from tests.support import (
    BaseTestCase,
    FileRegistry,
    InMemoryExcelStorage,
    get_fill_color,
    load_binary_excel_to_workbook,
)
from tests.support.contract_models import MergedContractImporter, SimpleContractImporter, creator, failing_creator


class TestImportContracts(BaseTestCase):
    @staticmethod
    def _build_workbook_bytes(*, sheet_name: str = 'Sheet1', rows: list[list[str | None]]) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = sheet_name

        for row_index, row in enumerate(rows, start=1):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    async def test_preflight_import_returns_valid_result_for_valid_workbook(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = alchemy.preflight_import(FileRegistry.TEST_SIMPLE_IMPORT)

        assert result.status == ImportPreflightStatus.VALID
        assert result.sheet_name == 'Sheet1'
        assert result.sheet_exists is True
        assert result.has_merged_header is False
        assert result.estimated_row_count == 1
        assert result.structural_issue_codes == []

    async def test_preflight_import_returns_header_invalid_for_invalid_header(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = alchemy.preflight_import(FileRegistry.TEST_HEADER_INVALID_INPUT)

        assert result.status == ImportPreflightStatus.HEADER_INVALID
        assert result.sheet_exists is True
        assert result.estimated_row_count == 1
        assert set(result.unrecognized) == {'不存在的表头'}
        assert '年龄' in set(result.missing_required)
        assert result.missing_primary == []

    async def test_preflight_import_reports_missing_primary_fields_in_update_mode(self):
        class UpdatePrimaryKeyImporter(BaseModel):
            employee_id: String = FieldMeta(label='员工编号', order=1, is_primary_key=True)
            name: String = FieldMeta(label='姓名', order=2)

        workbook_bytes = self._build_workbook_bytes(
            rows=[
                ['ignored hint'],
                ['姓名'],
                ['张三'],
            ]
        )
        storage = InMemoryExcelStorage(fixtures={'preflight-update-missing-primary.xlsx': workbook_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_update(UpdatePrimaryKeyImporter, storage=storage))

        result = alchemy.preflight_import('preflight-update-missing-primary.xlsx')

        assert result.status == ImportPreflightStatus.HEADER_INVALID
        assert result.sheet_exists is True
        assert result.estimated_row_count == 1
        assert result.missing_primary == ['员工编号']
        assert result.missing_required == []
        assert result.unrecognized == []
        assert result.duplicated == []

    async def test_preflight_import_reports_extra_fields_without_masking_present_required_fields(self):
        workbook_bytes = self._build_workbook_bytes(
            rows=[
                ['ignored hint'],
                [
                    '年龄',
                    '姓名',
                    '地址',
                    '是否启用',
                    '出生日期',
                    '邮箱',
                    '价格',
                    '爱好',
                    '公司',
                    '经理',
                    '部门',
                    '电话',
                    '单选',
                    '老板',
                    '领导',
                    '团队',
                    '网址',
                    '额外列',
                ],
                [
                    '18',
                    '张三',
                    '北京市',
                    '是',
                    '2021-01-01',
                    'noreply@example.com',
                    '100',
                    '篮球',
                    '阿里巴巴',
                    '李四',
                    '研发部',
                    '13800138000',
                    '选项1',
                    '马云',
                    '张三',
                    '研发部',
                    'https://www.baidu.com',
                    'unexpected',
                ],
            ]
        )
        storage = InMemoryExcelStorage(fixtures={'preflight-extra-field.xlsx': workbook_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        result = alchemy.preflight_import('preflight-extra-field.xlsx')

        assert result.status == ImportPreflightStatus.HEADER_INVALID
        assert result.sheet_exists is True
        assert result.estimated_row_count == 1
        assert result.missing_required == []
        assert result.missing_primary == []
        assert result.unrecognized == ['额外列']
        assert result.duplicated == []

    async def test_preflight_import_returns_sheet_missing_when_target_sheet_is_absent(self):
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = 'OtherSheet'
        worksheet['A1'] = 'ignored hint'
        worksheet['A2'] = '年龄'
        worksheet['B2'] = '姓名'

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()

        input_name = 'contract-preflight-missing-sheet.xlsx'
        buffer.seek(0)
        self.minio.put_object(self.minio.bucket_name, input_name, buffer, len(buffer.getvalue()))

        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = alchemy.preflight_import(input_name)

        assert result.status == ImportPreflightStatus.SHEET_MISSING
        assert result.sheet_name == 'Sheet1'
        assert result.sheet_exists is False
        assert result.has_merged_header is None
        assert result.estimated_row_count == 0
        assert result.structural_issue_codes == []

    async def test_preflight_import_returns_sheet_missing_for_explicit_storage_sheet_not_found_error(self):
        class MissingSheetStorage(InMemoryExcelStorage):
            def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str):
                raise WorksheetNotFoundError(
                    msg(MessageKey.WORKSHEET_NOT_FOUND, sheet_name=sheet_name),
                    message_key=MessageKey.WORKSHEET_NOT_FOUND,
                    sheet_name=sheet_name,
                )

        storage = MissingSheetStorage(fixtures={'ignored.xlsx': b'ignored'})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        result = alchemy.preflight_import('ignored.xlsx')

        assert result.status == ImportPreflightStatus.SHEET_MISSING
        assert result.sheet_name == 'Sheet1'
        assert result.sheet_exists is False
        assert result.structural_issue_codes == []

    async def test_preflight_import_returns_structure_invalid_for_missing_header_row(self):
        workbook_bytes = self._build_workbook_bytes(rows=[['ignored hint']])
        storage = InMemoryExcelStorage(fixtures={'header-missing.xlsx': workbook_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        result = alchemy.preflight_import('header-missing.xlsx')

        assert result.status == ImportPreflightStatus.STRUCTURE_INVALID
        assert result.sheet_name == 'Sheet1'
        assert result.sheet_exists is True
        assert result.has_merged_header is None
        assert result.estimated_row_count == 0
        assert result.structural_issue_codes == ['header_row_missing']

    async def test_preflight_import_reraises_unreadable_workbook_errors(self):
        storage = InMemoryExcelStorage(fixtures={'broken.xlsx': b'not-a-valid-workbook'})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        with self.assertRaisesRegex(Exception, 'File is not a zip file'):
            alchemy.preflight_import('broken.xlsx')

    async def test_preflight_import_reraises_unexpected_storage_errors(self):
        class ExplodingPreflightStorage(InMemoryExcelStorage):
            def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str):
                raise RuntimeError('boom')

        storage = ExplodingPreflightStorage(fixtures={'ignored.xlsx': b'ignored'})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        with self.assertRaisesRegex(RuntimeError, 'boom'):
            alchemy.preflight_import('ignored.xlsx')

    async def test_preflight_import_estimates_rows_for_merged_header_workbook(self):
        alchemy = ExcelAlchemy(ImporterConfig(MergedContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = alchemy.preflight_import(FileRegistry.TEST_IMPORT_WITH_MERGE_HEADER)

        assert result.status == ImportPreflightStatus.VALID
        assert result.sheet_exists is True
        assert result.has_merged_header is True
        assert result.estimated_row_count == 1

    async def test_preflight_import_estimates_rows_for_simple_header_with_multiple_data_rows(self):
        workbook_bytes = self._build_workbook_bytes(
            rows=[
                ['ignored hint'],
                [
                    '年龄',
                    '姓名',
                    '地址',
                    '是否启用',
                    '出生日期',
                    '邮箱',
                    '价格',
                    '爱好',
                    '公司',
                    '经理',
                    '部门',
                    '电话',
                    '单选',
                    '老板',
                    '领导',
                    '团队',
                    '网址',
                ],
                [
                    '18',
                    '张三',
                    '北京市',
                    '是',
                    '2021-01-01',
                    'noreply@example.com',
                    '100',
                    '篮球',
                    '阿里巴巴',
                    '李四',
                    '研发部',
                    '13800138000',
                    '选项1',
                    '马云',
                    '张三',
                    '研发部',
                    'https://www.baidu.com',
                ],
                [
                    '19',
                    '李四',
                    '上海市',
                    '否',
                    '2021-01-02',
                    'person@example.com',
                    '200',
                    '足球',
                    '腾讯',
                    '王五',
                    '市场部',
                    '13900139000',
                    '选项2',
                    '马化腾',
                    '李四',
                    '市场部',
                    'https://example.com',
                ],
            ]
        )
        storage = InMemoryExcelStorage(fixtures={'preflight-two-rows.xlsx': workbook_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        result = alchemy.preflight_import('preflight-two-rows.xlsx')

        assert result.status == ImportPreflightStatus.VALID
        assert result.sheet_exists is True
        assert result.has_merged_header is False
        assert result.estimated_row_count == 2

    async def test_preflight_import_does_not_execute_row_callbacks_or_mutate_last_import_session(self):
        context: dict[str, object] = {'created_rows': []}

        async def tracking_creator(
            data: dict[str, object], runtime_context: dict[str, object] | None
        ) -> dict[str, object]:
            assert runtime_context is not None
            created_rows = runtime_context.setdefault('created_rows', [])
            assert isinstance(created_rows, list)
            created_rows.append(data.copy())
            return data

        alchemy = ExcelAlchemy(
            ImporterConfig.for_create(SimpleContractImporter, creator=tracking_creator, minio=cast(Minio, self.minio))
        )
        alchemy.add_context(context)

        result = alchemy.preflight_import(FileRegistry.TEST_SIMPLE_IMPORT)

        assert result.status == ImportPreflightStatus.VALID
        assert context['created_rows'] == []
        assert alchemy.last_import_snapshot is None

    async def test_preflight_import_does_not_upload_or_populate_error_maps(self):
        workbook_bytes = self._build_workbook_bytes(
            rows=[
                ['ignored hint'],
                [
                    '年龄',
                    '姓名',
                    '地址',
                    '是否启用',
                    '出生日期',
                    '邮箱',
                    '价格',
                    '爱好',
                    '公司',
                    '经理',
                    '部门',
                    '电话',
                    '单选',
                    '老板',
                    '领导',
                    '团队',
                    '网址',
                ],
                [
                    '18',
                    '张三',
                    '北京市',
                    '是',
                    '2021-01-01',
                    'noreply@example.com',
                    '100',
                    '篮球',
                    '阿里巴巴',
                    '李四',
                    '研发部',
                    '13800138000',
                    '选项1',
                    '马云',
                    '张三',
                    '研发部',
                    'https://www.baidu.com',
                ],
            ]
        )
        storage = InMemoryExcelStorage(fixtures={'preflight-no-side-effects.xlsx': workbook_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))

        result = alchemy.preflight_import('preflight-no-side-effects.xlsx')

        assert result.status == ImportPreflightStatus.VALID
        assert storage.uploaded == {}
        assert alchemy.last_import_snapshot is None
        assert alchemy.cell_error_map == {}
        assert alchemy.row_error_map == {}

    async def test_import_data_emits_expected_success_events(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))
        events: list[dict[str, object]] = []

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
            output_excel_name='contract-success-events.xlsx',
            on_event=events.append,
        )

        assert result.result == ValidateResult.SUCCESS
        assert [event['event'] for event in events] == [
            'started',
            'header_validated',
            'row_processed',
            'completed',
        ]
        assert events[1] == {
            'event': 'header_validated',
            'is_valid': True,
        }
        assert events[2] == {
            'event': 'row_processed',
            'processed_row_count': 1,
            'total_row_count': 1,
            'success_count': 1,
            'fail_count': 0,
        }
        assert events[3] == {
            'event': 'completed',
            'result': 'SUCCESS',
            'success_count': 1,
            'fail_count': 0,
            'url': None,
        }

    async def test_import_data_returns_success_result_for_valid_workbook(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
            output_excel_name='contract-success.xlsx',
        )

        assert result.result == ValidateResult.SUCCESS
        assert result.success_count == 1
        assert result.fail_count == 0
        assert result.url is None

    async def test_import_data_returns_header_invalid_result_for_invalid_header(self):
        output_name = 'contract-header-invalid.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_HEADER_INVALID_INPUT,
            output_excel_name=output_name,
        )

        assert result.result == ValidateResult.HEADER_INVALID
        assert set(result.unrecognized) == {'不存在的表头'}
        assert '年龄' in set(result.missing_required)
        assert output_name not in self.minio.storage

    async def test_import_data_emits_expected_header_invalid_events(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))
        events: list[dict[str, object]] = []

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_HEADER_INVALID_INPUT,
            output_excel_name='contract-header-invalid-events.xlsx',
            on_event=events.append,
        )

        assert result.result == ValidateResult.HEADER_INVALID
        assert [event['event'] for event in events] == [
            'started',
            'header_validated',
            'completed',
        ]
        assert events[1]['event'] == 'header_validated'
        assert events[1]['is_valid'] is False
        missing_required = events[1]['missing_required']
        missing_primary = events[1]['missing_primary']
        unrecognized = events[1]['unrecognized']
        duplicated = events[1]['duplicated']
        assert isinstance(missing_required, list)
        assert isinstance(missing_primary, list)
        assert isinstance(unrecognized, list)
        assert isinstance(duplicated, list)
        assert '年龄' in missing_required
        assert missing_primary == []
        assert unrecognized == ['不存在的表头']
        assert duplicated == []
        assert events[2] == {
            'event': 'completed',
            'result': 'HEADER_INVALID',
            'success_count': 0,
            'fail_count': 0,
            'url': None,
        }

    async def test_import_data_reloads_workbook_state_on_each_run(self):
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        first_result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_HEADER_INVALID_INPUT,
            output_excel_name='contract-first-header-invalid.xlsx',
        )
        second_result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
            output_excel_name='contract-second-success.xlsx',
        )

        assert first_result.result == ValidateResult.HEADER_INVALID
        assert second_result.result == ValidateResult.SUCCESS
        assert second_result.success_count == 1
        assert second_result.fail_count == 0
        assert second_result.url is None

    async def test_import_session_snapshot_tracks_completed_successful_run(self):
        alchemy = ExcelAlchemy(
            ImporterConfig.for_create(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio))
        )

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
            output_excel_name='contract-session-success.xlsx',
        )

        snapshot = alchemy.last_import_snapshot

        assert result.result == ValidateResult.SUCCESS
        assert snapshot is not None
        assert snapshot.phase == ImportSessionPhase.COMPLETED
        assert snapshot.result == ValidateResult.SUCCESS
        assert snapshot.data_row_count == 1
        assert snapshot.processed_row_count == 1
        assert snapshot.success_count == 1
        assert snapshot.fail_count == 0
        assert not snapshot.rendered_result_workbook

    async def test_import_data_uploads_result_workbook_for_invalid_rows(self):
        output_name = 'contract-data-invalid.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR,
            output_excel_name=output_name,
        )

        assert result.result == ValidateResult.DATA_INVALID
        assert result.success_count == 0
        assert result.fail_count == 1
        assert result.url == f'excel/{output_name}'
        assert output_name in self.minio.storage

    async def test_import_data_emits_expected_data_invalid_events(self):
        output_name = 'contract-data-invalid-events.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))
        events: list[dict[str, object]] = []

        result = await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR,
            output_excel_name=output_name,
            on_event=events.append,
        )

        assert result.result == ValidateResult.DATA_INVALID
        assert [event['event'] for event in events] == [
            'started',
            'header_validated',
            'row_processed',
            'completed',
        ]
        assert events[2] == {
            'event': 'row_processed',
            'processed_row_count': 1,
            'total_row_count': 1,
            'success_count': 0,
            'fail_count': 1,
        }
        assert events[3] == {
            'event': 'completed',
            'result': 'DATA_INVALID',
            'success_count': 0,
            'fail_count': 1,
            'url': f'excel/{output_name}',
        }

    async def test_import_data_emits_failed_event_for_unexpected_exception(self):
        class ExplodingReadStorage(InMemoryExcelStorage):
            def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str):
                raise RuntimeError('boom')

        source_bytes = self.minio.storage[FileRegistry.TEST_SIMPLE_IMPORT]['data'].getvalue()
        storage = ExplodingReadStorage(fixtures={FileRegistry.TEST_SIMPLE_IMPORT: source_bytes})
        alchemy = ExcelAlchemy(ImporterConfig.for_create(SimpleContractImporter, creator=creator, storage=storage))
        events: list[dict[str, object]] = []

        with self.assertRaisesRegex(RuntimeError, 'boom'):
            await alchemy.import_data(
                input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
                output_excel_name='contract-failed-events.xlsx',
                on_event=events.append,
            )

        assert [event['event'] for event in events] == [
            'started',
            'failed',
        ]
        assert events[1] == {
            'event': 'failed',
            'error_type': 'RuntimeError',
            'error_message': 'boom',
        }

    async def test_import_result_workbook_returns_result_and_reason_columns(self):
        output_name = 'contract-data-invalid-columns.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR,
            output_excel_name=output_name,
        )
        workbook = load_binary_excel_to_workbook(self.minio.storage[output_name]['data'].getvalue())
        worksheet = workbook['Sheet1']

        assert worksheet['A2'].value == RESULT_COLUMN_LABEL
        assert worksheet['B2'].value == REASON_COLUMN_LABEL
        assert worksheet['A3'].value == '校验不通过'
        assert isinstance(worksheet['B3'].value, str)
        assert worksheet['B3'].value.startswith('1、')
        assert '【出生日期】' in worksheet['B3'].value

    async def test_import_result_workbook_marks_failed_cells_in_red(self):
        output_name = 'contract-data-invalid-colors.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR,
            output_excel_name=output_name,
        )
        workbook = load_binary_excel_to_workbook(self.minio.storage[output_name]['data'].getvalue())
        worksheet = workbook['Sheet1']
        row_colors = [get_fill_color(cell) for cell in worksheet[3]]

        assert BACKGROUND_ERROR_COLOR in row_colors

    async def test_import_result_workbook_marks_business_cell_errors_in_red(self):
        output_name = 'contract-data-invalid-business-cell.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(
            ImporterConfig(SimpleContractImporter, creator=failing_creator, minio=cast(Minio, self.minio))
        )

        await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT,
            output_excel_name=output_name,
        )
        workbook = load_binary_excel_to_workbook(self.minio.storage[output_name]['data'].getvalue())
        worksheet = workbook['Sheet1']

        assert worksheet['B3'].value == '1、【姓名】Simulated failure'
        assert get_fill_color(worksheet['D3']) == BACKGROUND_ERROR_COLOR

    async def test_import_result_workbook_supports_english_display_locale(self):
        output_name = 'contract-data-invalid-english.xlsx'
        self.minio.storage.pop(output_name, None)
        alchemy = ExcelAlchemy(
            ImporterConfig(SimpleContractImporter, creator=creator, minio=cast(Minio, self.minio), locale='en')
        )

        await alchemy.import_data(
            input_excel_name=FileRegistry.TEST_SIMPLE_IMPORT_WITH_ERROR,
            output_excel_name=output_name,
        )
        workbook = load_binary_excel_to_workbook(self.minio.storage[output_name]['data'].getvalue())
        worksheet = workbook['Sheet1']

        assert worksheet['A2'].value == 'Validation result\nDelete this column before re-uploading'
        assert worksheet['B2'].value == 'Failure reason\nDelete this column before re-uploading'
        assert worksheet['A3'].value == 'Validation failed'

    async def test_import_result_workbook_marks_merged_header_failures_on_the_correct_data_row(self):
        input_name = 'contract-merged-invalid-input.xlsx'
        output_name = 'contract-merged-invalid-output.xlsx'
        self.minio.storage.pop(output_name, None)

        source_content = self.minio.storage[FileRegistry.TEST_IMPORT_WITH_MERGE_HEADER]['data'].getvalue()
        workbook = load_workbook(io.BytesIO(source_content))
        worksheet = workbook['Sheet1']
        worksheet['E4'] = 'not-a-date'

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)
        self.minio.put_object(self.minio.bucket_name, input_name, buffer, len(buffer.getvalue()))

        alchemy = ExcelAlchemy(ImporterConfig(MergedContractImporter, creator=creator, minio=cast(Minio, self.minio)))

        result = await alchemy.import_data(
            input_excel_name=input_name,
            output_excel_name=output_name,
        )

        assert result.result == ValidateResult.DATA_INVALID

        result_workbook = load_binary_excel_to_workbook(self.minio.storage[output_name]['data'].getvalue())
        result_worksheet = result_workbook['Sheet1']

        assert result_worksheet['A4'].value == '校验不通过'
        assert isinstance(result_worksheet['B4'].value, str)
        assert '【出生日期】' in result_worksheet['B4'].value
