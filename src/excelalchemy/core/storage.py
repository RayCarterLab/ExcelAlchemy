"""Storage adapters used by ExcelAlchemy to read and upload workbooks."""

from typing import BinaryIO, cast

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excelalchemy.core.table import WorksheetTable
from excelalchemy.exc import ConfigError
from excelalchemy.types.alchemy import ExporterConfig, ImporterConfig
from excelalchemy.types.identity import UrlStr
from excelalchemy.util.file import read_file_from_minio_object, remove_excel_prefix, upload_file_from_minio_object


class MinioStorageGateway:
    """Small gateway around the Minio-backed workbook IO helpers."""

    def __init__(self, config: ImporterConfig | ExporterConfig):
        self.config = config

    def read_excel_dataframe(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        """Read one workbook object from Minio into a worksheet table."""
        if self.config.minio is None:
            raise ConfigError('未配置 minio')

        file_object = read_file_from_minio_object(
            self.config.minio,
            self.config.bucket_name,
            input_excel_name,
        )

        try:
            file_object.seek(0)
            workbook = load_workbook(cast(BinaryIO, file_object), data_only=True)
            try:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f'Worksheet named {sheet_name!r} not found')
                worksheet = workbook[sheet_name]
                return self._worksheet_to_table(worksheet, skiprows=skiprows)
            finally:
                workbook.close()
        finally:
            cast(BinaryIO, file_object).close()

    def _worksheet_to_table(self, worksheet: Worksheet, *, skiprows: int) -> WorksheetTable:
        rows = [
            [self._normalize_cell_value(value) for value in row]
            for row in worksheet.iter_rows(
                min_row=skiprows + 1,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
                values_only=True,
            )
        ]

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    @staticmethod
    def _normalize_cell_value(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            return value
        return str(value)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        """Upload one rendered workbook and return its signed URL."""
        if self.config.minio is None:
            raise ConfigError('未配置 minio')
        url = upload_file_from_minio_object(
            self.config.minio,
            self.config.bucket_name,
            output_name,
            remove_excel_prefix(content_with_prefix),
            self.config.url_expires,
        )
        return UrlStr(url)
