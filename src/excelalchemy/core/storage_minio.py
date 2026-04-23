"""Minio-backed Excel storage implementation."""

import base64
import io
from datetime import timedelta
from typing import IO, BinaryIO, cast

from minio import Minio
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel
from urllib3.response import BaseHTTPResponse

from excelalchemy._primitives.identity import UrlStr
from excelalchemy.config import ExporterConfig, ImporterConfig
from excelalchemy.core.storage_protocol import ExcelStorage
from excelalchemy.core.table import WorksheetTable
from excelalchemy.exceptions import ConfigError, WorksheetNotFoundError
from excelalchemy.i18n.messages import MessageKey
from excelalchemy.i18n.messages import message as msg
from excelalchemy.util.file import remove_excel_prefix


class MinioStorageGateway(ExcelStorage):
    """Excel storage strategy backed by a Minio-compatible object store."""

    def __init__[
        ContextT,
        ImportCreateModelT: BaseModel,
        ImportUpdateModelT: BaseModel,
        ExportModelT: BaseModel,
    ](
        self,
        config: ImporterConfig[ContextT, ImportCreateModelT, ImportUpdateModelT] | ExporterConfig[ExportModelT],
    ):
        self.config = config
        self.storage_options = config.storage_options

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        """Read one workbook object from Minio into a worksheet table."""
        if self.storage_options.minio is None:
            raise ConfigError(msg(MessageKey.MINIO_CLIENT_NOT_CONFIGURED))

        file_object = self._read_file_object(
            self.storage_options.minio,
            self.storage_options.bucket_name,
            input_excel_name,
        )

        try:
            file_object.seek(0)
            workbook = load_workbook(cast(BinaryIO, file_object), data_only=True)
            try:
                if sheet_name not in workbook.sheetnames:
                    raise WorksheetNotFoundError(
                        msg(MessageKey.WORKSHEET_NOT_FOUND, sheet_name=sheet_name),
                        message_key=MessageKey.WORKSHEET_NOT_FOUND,
                        sheet_name=sheet_name,
                    )
                worksheet = workbook[sheet_name]
                return self._worksheet_to_table(worksheet, skiprows=skiprows)
            finally:
                workbook.close()
        finally:
            cast(BinaryIO, file_object).close()

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        """Upload one rendered workbook and return its signed URL."""
        if self.storage_options.minio is None:
            raise ConfigError(msg(MessageKey.MINIO_CLIENT_NOT_CONFIGURED))
        url = self._upload_file_object(
            self.storage_options.minio,
            self.storage_options.bucket_name,
            output_name,
            remove_excel_prefix(content_with_prefix),
            self.storage_options.url_expires,
        )
        return UrlStr(url)

    def read_excel_dataframe(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        """Backward-compatible alias for the worksheet-table reader."""
        return self.read_excel_table(input_excel_name, skiprows=skiprows, sheet_name=sheet_name)

    def _worksheet_to_table(self, worksheet: Worksheet, *, skiprows: int) -> WorksheetTable:
        rows = [
            [self._normalize_cell_value(value) for value in row]
            for row in worksheet.iter_rows(
                min_row=skiprows + 1,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
                values_only=True,
            )
        ]

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    @staticmethod
    def _normalize_cell_value(value: object) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            return value
        return str(value)

    @staticmethod
    def _construct_file_like_object(response: BaseHTTPResponse) -> IO[bytes]:
        """Construct a file-like object from an object storage response."""
        return io.BytesIO(response.read())

    @classmethod
    def _read_file_object(cls, client: Minio, bucket_name: str, filename: str) -> IO[bytes]:
        response: BaseHTTPResponse = client.get_object(bucket_name, filename)
        return cls._construct_file_like_object(response)

    @staticmethod
    def _upload_file_object(
        client: Minio,
        bucket_name: str,
        filename: str,
        content: str,
        expires: int,
    ) -> str:
        data = base64.b64decode(content)
        client.put_object(bucket_name, filename, io.BytesIO(data), len(data))
        return client.presigned_get_object(
            bucket_name,
            filename,
            expires=timedelta(seconds=expires),
        )
