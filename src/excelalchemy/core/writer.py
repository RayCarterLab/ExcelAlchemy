"""Render Excel workbooks with openpyxl only."""

import base64
import io
from collections import defaultdict
from math import ceil
from typing import Any, BinaryIO, cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excelalchemy._primitives.constants import (
    BACKGROUND_ERROR_COLOR,
    BACKGROUND_REQUIRED_COLOR,
    CHARACTER_WIDTH,
    DEFAULT_SHEET_NAME,
    FONT_READ_COLOR,
)
from excelalchemy._primitives.identity import ColumnIndex, DataUrlStr, Label, RowIndex, UniqueLabel
from excelalchemy.core.table import WorksheetTable
from excelalchemy.exceptions import ExcelCellError
from excelalchemy.i18n.messages import MessageKey
from excelalchemy.i18n.messages import display_message as dmsg
from excelalchemy.i18n.messages import message as msg
from excelalchemy.metadata import FieldMetaInfo
from excelalchemy.results import ValidateRowResult
from excelalchemy.util.file import add_excel_prefix, value_is_nan

OPENPYXL_EXCEL_INDEX_START_AT = 1

HEADER_HINT_ROW_INDEX = 1
HEADER_HINT_COL_INDEX = 1
HEADER_HINT_LINE_COUNT = 1

SIMPLE_HEADER_ROW_COUNT = 1
MERGE_HEADER_ROW_COUNT = 2


def _get_file(file: BinaryIO | None = None) -> BinaryIO:
    """Return the writable buffer used to build the workbook payload."""
    return cast(BinaryIO, file or io.BytesIO())


def _create_workbook(sheet_name: str) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet_name
    return workbook, worksheet


def _worksheet_cell(worksheet: Worksheet, *, row: int, column: int) -> Cell:
    return cast(Cell, worksheet.cell(row=row, column=column))


def _encode_workbook(workbook: Workbook, file: BinaryIO, *, close_file: bool) -> DataUrlStr:
    workbook.save(file)
    file.seek(0)
    content = base64.b64encode(file.read()).decode()
    if close_file:
        file.close()
    return DataUrlStr(add_excel_prefix(content))


def _build_comment(field_meta: FieldMetaInfo) -> Comment | None:
    comment_text = field_meta.excel_codec.build_comment(field_meta)
    if not comment_text:
        return None

    return Comment(
        text=comment_text,
        author='https://github.com/SundayWindy/ExcelAlchemy',
        height=sum(ceil(len(line) / 20) for line in comment_text.splitlines()) * 28,
        width=300,
    )


def _style_header_cell(cell: Cell, field_meta: FieldMetaInfo) -> None:
    comment = _build_comment(field_meta)
    if comment is not None:
        cell.comment = comment
    if field_meta.required:
        cell.fill = PatternFill(start_color=BACKGROUND_REQUIRED_COLOR, fill_type='solid')
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.number_format = numbers.FORMAT_TEXT


def _style_child_header_cell(cell: Cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.number_format = numbers.FORMAT_TEXT


def _write_header_hint(worksheet: Worksheet, *, column_count: int) -> None:
    cell = _worksheet_cell(worksheet, row=HEADER_HINT_ROW_INDEX, column=HEADER_HINT_COL_INDEX)
    cell.value = dmsg(MessageKey.HEADER_HINT)
    cell.font = Font(size=16)
    cell.alignment = Alignment(wrap_text=True)
    worksheet.merge_cells(
        start_row=HEADER_HINT_ROW_INDEX,
        start_column=HEADER_HINT_COL_INDEX,
        end_row=HEADER_HINT_ROW_INDEX,
        end_column=max(column_count, HEADER_HINT_COL_INDEX),
    )
    worksheet.row_dimensions[HEADER_HINT_ROW_INDEX].height = 120


def _write_simple_header(
    worksheet: Worksheet,
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    column_write_offset: int = 0,
    row_write_offset: int = 0,
) -> None:
    header_row_index = row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT

    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):
        field_meta = field_meta_mapping[cast(UniqueLabel, column)]
        cell = _worksheet_cell(worksheet, row=header_row_index, column=openpyxl_col_index)
        cell.value = str(column)
        _style_header_cell(cell, field_meta)


def _write_vertically_merged_header(
    worksheet: Worksheet,
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    start_row: int,
    column_write_offset: int,
) -> None:
    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):
        field_meta = field_meta_mapping[cast(UniqueLabel, column)]
        if field_meta.label == field_meta.parent_label:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=openpyxl_col_index,
                end_row=start_row + 1,
                end_column=openpyxl_col_index,
            )


def _write_horizontally_merged_header(
    worksheet: Worksheet,
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    start_row: int,
    column_write_offset: int,
) -> None:
    counter: dict[Label, int] = defaultdict(int)
    for field_meta in field_meta_mapping.values():
        if field_meta.parent_label is None:
            raise RuntimeError(msg(MessageKey.PARENT_LABEL_EMPTY_RUNTIME))
        counter[field_meta.parent_label] += 1

    for openpyxl_col_index, column in enumerate(
        df.columns[column_write_offset:],
        start=column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT,
    ):
        field_meta = field_meta_mapping[cast(UniqueLabel, column)]
        if field_meta.parent_label is None:
            raise RuntimeError(msg(MessageKey.PARENT_LABEL_EMPTY_RUNTIME))
        if field_meta.label != field_meta.parent_label and field_meta.offset == 0:
            cell = _worksheet_cell(worksheet, row=start_row, column=openpyxl_col_index)
            cell.value = str(field_meta.parent_label)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            worksheet.merge_cells(
                start_row=start_row,
                start_column=openpyxl_col_index,
                end_row=start_row,
                end_column=openpyxl_col_index + counter[field_meta.parent_label] - 1,
            )


def _write_merged_header(
    worksheet: Worksheet,
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    column_write_offset: int = 0,
    row_write_offset: int = 0,
) -> None:
    _write_simple_header(
        worksheet,
        df,
        field_meta_mapping,
        column_write_offset=column_write_offset,
        row_write_offset=row_write_offset,
    )

    child_row_index = row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT + 1
    child_headers = df.iloc[0].tolist()
    for column_index, child_value in enumerate(child_headers, start=OPENPYXL_EXCEL_INDEX_START_AT):
        cell = _worksheet_cell(worksheet, row=child_row_index, column=column_index + column_write_offset)
        cell.value = str(child_value)
        _style_child_header_cell(cell)

    start_row = row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
    _write_vertically_merged_header(
        worksheet,
        df,
        field_meta_mapping,
        start_row=start_row,
        column_write_offset=column_write_offset,
    )
    _write_horizontally_merged_header(
        worksheet,
        df,
        field_meta_mapping,
        start_row=start_row,
        column_write_offset=column_write_offset,
    )


def _get_parsed_value(
    df: WorksheetTable,
    row_index: int,
    col_index: int,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
) -> str:
    cell_value: str | Any | None = df.iloc[row_index, col_index]

    if value_is_nan(cell_value):
        return ''

    col_label = cast(UniqueLabel, df.columns[col_index])
    if col_label not in field_meta_mapping:
        return str(cell_value)

    field_meta = field_meta_mapping[col_label]
    parsed_value = field_meta.excel_codec.format_display_value(cell_value, field_meta)
    return str(parsed_value)


def _mark_error(
    worksheet: Worksheet,
    errors: dict[RowIndex, dict[ColumnIndex, list[ExcelCellError]]],
    *,
    column_write_offset: int,
    row_write_offset: int,
) -> None:
    for row_index, cols in errors.items():
        for col_index, exceptions in cols.items():
            if not exceptions:
                continue

            openpyxl_col_index = col_index + column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
            openpyxl_row_index = row_index + row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
            cell = _worksheet_cell(worksheet, row=openpyxl_row_index, column=openpyxl_col_index)
            cell.fill = PatternFill(
                start_color=BACKGROUND_ERROR_COLOR,
                end_color=BACKGROUND_ERROR_COLOR,
                fill_type='solid',
            )
            cell.alignment = Alignment(wrap_text=True)


def _write_value(
    df: WorksheetTable,
    worksheet: Worksheet,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    table_data_start_index: int,
    column_write_offset: int,
    row_write_offset: int,
) -> None:
    col_width_mapping: dict[ColumnIndex, float] = defaultdict(float)

    for row_index in range(table_data_start_index, df.shape[0]):
        for column_index in range(df.shape[1]):
            openpyxl_col_index = column_index + column_write_offset + OPENPYXL_EXCEL_INDEX_START_AT
            openpyxl_row_index = row_index + row_write_offset + OPENPYXL_EXCEL_INDEX_START_AT

            cell = _worksheet_cell(worksheet, row=openpyxl_row_index, column=openpyxl_col_index)
            cell.value = _get_parsed_value(df, row_index, column_index, field_meta_mapping)
            cell.number_format = numbers.FORMAT_TEXT
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            if dmsg(MessageKey.RESULT_COLUMN_LABEL) == df.columns[column_index] and cell.value == str(ValidateRowResult.FAIL):
                cell.font = Font(color=FONT_READ_COLOR)

            col_width_mapping[ColumnIndex(openpyxl_col_index)] = max(
                col_width_mapping[ColumnIndex(openpyxl_col_index)],
                max(len(str(part)) for part in str(cell.value).split('\n')),
                len(str(df.columns[column_index])),
            )

    for openpyxl_col_index, width in col_width_mapping.items():
        worksheet.column_dimensions[get_column_letter(openpyxl_col_index)].width = round(
            (width + 4) * CHARACTER_WIDTH, 2
        )


def _write_value_mark_error(
    worksheet: Worksheet,
    df: WorksheetTable,
    errors: dict[RowIndex, dict[ColumnIndex, list[ExcelCellError]]],
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    *,
    row_write_offset: int = 0,
    column_write_offset: int = 0,
    table_data_start_index: int = 0,
) -> None:
    _mark_error(
        worksheet,
        errors,
        column_write_offset=column_write_offset,
        row_write_offset=row_write_offset,
    )
    _write_value(
        df,
        worksheet,
        field_meta_mapping,
        table_data_start_index=table_data_start_index,
        row_write_offset=row_write_offset,
        column_write_offset=column_write_offset,
    )


def render_simple_header_excel(
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    column_write_offset: int = 0,
) -> DataUrlStr:
    if file is None:
        close_file = True

    tmp = _get_file(file)
    workbook, worksheet = _create_workbook(sheet_name)
    _write_header_hint(worksheet, column_count=len(df.columns))
    _write_simple_header(
        worksheet,
        df,
        field_meta_mapping,
        column_write_offset=column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT,
    )
    _write_value(
        df,
        worksheet,
        field_meta_mapping,
        table_data_start_index=0,
        column_write_offset=column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT + SIMPLE_HEADER_ROW_COUNT,
    )

    return _encode_workbook(workbook, tmp, close_file=close_file)


def render_merged_header_excel(
    df: WorksheetTable,
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    column_write_offset: int = 0,
) -> DataUrlStr:
    if file is None:
        close_file = True

    tmp = _get_file(file)
    workbook, worksheet = _create_workbook(sheet_name)
    _write_header_hint(worksheet, column_count=len(df.columns))
    _write_merged_header(
        worksheet,
        df,
        field_meta_mapping,
        column_write_offset=column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT,
    )
    _write_value(
        df,
        worksheet,
        field_meta_mapping,
        table_data_start_index=1,
        column_write_offset=column_write_offset,
        row_write_offset=HEADER_HINT_LINE_COUNT + SIMPLE_HEADER_ROW_COUNT,
    )

    return _encode_workbook(workbook, tmp, close_file=close_file)


def render_data_excel(
    df: WorksheetTable,
    errors: dict[RowIndex, dict[ColumnIndex, list[ExcelCellError]]],
    field_meta_mapping: dict[UniqueLabel, FieldMetaInfo],
    sheet_name: str = DEFAULT_SHEET_NAME,
    file: BinaryIO | None = None,
    close_file: bool = True,
    has_merged_header: bool = False,
) -> DataUrlStr:
    if file is None:
        close_file = True

    tmp = _get_file(file)
    workbook, worksheet = _create_workbook(sheet_name)
    _write_header_hint(worksheet, column_count=len(df.columns))

    if has_merged_header:
        table_data_start_index = 1
        _write_merged_header(
            worksheet,
            df,
            field_meta_mapping,
            row_write_offset=HEADER_HINT_LINE_COUNT,
        )
    else:
        table_data_start_index = 0
        _write_simple_header(
            worksheet,
            df,
            field_meta_mapping,
            row_write_offset=HEADER_HINT_LINE_COUNT,
        )

    _write_value_mark_error(
        worksheet,
        df,
        errors,
        field_meta_mapping,
        row_write_offset=HEADER_HINT_LINE_COUNT + SIMPLE_HEADER_ROW_COUNT,
        table_data_start_index=table_data_start_index,
    )

    return _encode_workbook(workbook, tmp, close_file=close_file)
