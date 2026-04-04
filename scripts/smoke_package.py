"""Release smoke test for installed ExcelAlchemy packages."""

from __future__ import annotations

import asyncio
import io
from base64 import b64decode

from openpyxl import load_workbook
from pydantic import BaseModel

from excelalchemy import ExcelAlchemy, ExcelStorage, ExporterConfig, FieldMeta, ImporterConfig, Number, String, UrlStr
from excelalchemy.core.table import WorksheetTable


class SmokeImporter(BaseModel):
    full_name: String = FieldMeta(label='Full name', order=1)
    age: Number = FieldMeta(label='Age', order=2)


class InMemorySmokeStorage(ExcelStorage):
    def __init__(self) -> None:
        self.fixtures: dict[str, bytes] = {}
        self.uploaded: dict[str, bytes] = {}

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        workbook = load_workbook(io.BytesIO(self.fixtures[input_excel_name]), data_only=True)
        try:
            worksheet = workbook[sheet_name]
            rows = [
                [None if value is None else str(value) for value in row]
                for row in worksheet.iter_rows(
                    min_row=skiprows + 1,
                    max_row=worksheet.max_row,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            ]
        finally:
            workbook.close()

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        _, payload = content_with_prefix.split(',', 1)
        self.uploaded[output_name] = b64decode(payload)
        return UrlStr(f'memory://{output_name}')


async def _create_employee(row: dict[str, object], context: object) -> dict[str, object]:
    return row


def _build_import_fixture(storage: InMemorySmokeStorage, template_bytes: bytes) -> None:
    workbook = load_workbook(io.BytesIO(template_bytes))
    try:
        worksheet = workbook['Sheet1']
        worksheet['A3'] = 'TaylorChen'
        worksheet['B3'] = '32'

        buffer = io.BytesIO()
        workbook.save(buffer)
        storage.fixtures['smoke-input.xlsx'] = buffer.getvalue()
    finally:
        workbook.close()


def _build_invalid_import_fixture(storage: InMemorySmokeStorage, template_bytes: bytes) -> None:
    workbook = load_workbook(io.BytesIO(template_bytes))
    try:
        worksheet = workbook['Sheet1']
        worksheet['A3'] = 'TaylorChen'
        worksheet['B3'] = 'invalid-age'

        buffer = io.BytesIO()
        workbook.save(buffer)
        storage.fixtures['smoke-invalid-input.xlsx'] = buffer.getvalue()
    finally:
        workbook.close()


async def main() -> None:
    storage = InMemorySmokeStorage()

    importer = ExcelAlchemy(
        ImporterConfig.for_create(
            SmokeImporter,
            creator=_create_employee,
            storage=storage,
            locale='en',
        )
    )
    template = importer.download_template_artifact(filename='smoke-template.xlsx')
    assert len(template.as_bytes()) > 0

    _build_import_fixture(storage, template.as_bytes())
    import_result = await importer.import_data('smoke-input.xlsx', 'smoke-result.xlsx')
    assert import_result.success_count == 1
    assert import_result.fail_count == 0
    assert import_result.result == 'SUCCESS'

    invalid_importer = ExcelAlchemy(
        ImporterConfig.for_create(
            SmokeImporter,
            creator=_create_employee,
            storage=storage,
            locale='en',
        )
    )
    _build_invalid_import_fixture(storage, template.as_bytes())
    invalid_result = await invalid_importer.import_data('smoke-invalid-input.xlsx', 'smoke-invalid-result.xlsx')
    assert invalid_result.result == 'DATA_INVALID'
    assert invalid_result.fail_count == 1
    assert invalid_importer.cell_error_map.error_count >= 1
    assert invalid_importer.row_error_map.error_count >= 1
    invalid_result_payload = invalid_result.to_api_payload()
    cell_payload = invalid_importer.cell_error_map.to_api_payload()
    row_payload = invalid_importer.row_error_map.to_api_payload()
    assert invalid_result_payload['result'] == 'DATA_INVALID'
    assert invalid_result_payload['is_data_invalid'] is True
    assert invalid_result_payload['summary']['fail_count'] == 1
    assert cell_payload['error_count'] >= 1
    assert row_payload['error_count'] >= 1
    assert isinstance(cell_payload['items'], list)
    assert isinstance(row_payload['items'], list)
    assert cell_payload['summary']['by_code']
    assert cell_payload['summary']['by_row']
    assert row_payload['summary']['by_code']
    assert row_payload['summary']['by_row']
    first_cell_issue = cell_payload['items'][0]
    assert isinstance(first_cell_issue['code'], str) and first_cell_issue['code']
    assert first_cell_issue['field_label'] == 'Age'
    assert isinstance(first_cell_issue['message'], str) and first_cell_issue['message']
    assert isinstance(first_cell_issue['display_message'], str) and first_cell_issue['display_message']
    assert first_cell_issue['row_number_for_humans'] == 1
    assert isinstance(first_cell_issue['column_number_for_humans'], int)
    assert first_cell_issue['column_number_for_humans'] >= 1
    first_row_issue = row_payload['items'][0]
    assert isinstance(first_row_issue['code'], str) and first_row_issue['code']
    assert isinstance(first_row_issue['message'], str) and first_row_issue['message']
    assert isinstance(first_row_issue['display_message'], str) and first_row_issue['display_message']
    assert first_row_issue['row_number_for_humans'] == 1

    exporter = ExcelAlchemy(ExporterConfig.for_storage(SmokeImporter, storage=storage, locale='en'))
    artifact = exporter.export_artifact(
        [{'full_name': 'TaylorChen', 'age': 32}],
        filename='smoke-export.xlsx',
    )
    assert len(artifact.as_bytes()) > 0
    uploaded_url = exporter.export_upload('smoke-export-upload.xlsx', [{'full_name': 'TaylorChen', 'age': 32}])
    assert uploaded_url == 'memory://smoke-export-upload.xlsx'
    assert 'smoke-export-upload.xlsx' in storage.uploaded

    print('Package smoke passed')
    print(f'Import result: {import_result.result}')
    print(f'Invalid import result: {invalid_result.result}')
    print(f'Upload URL: {uploaded_url}')


if __name__ == '__main__':
    asyncio.run(main())
