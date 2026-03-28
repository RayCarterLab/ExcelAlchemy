"""Generate English Excel assets for portfolio screenshots."""

from __future__ import annotations

import asyncio
import base64
import shutil
from pathlib import Path
from typing import Annotated

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from excelalchemy import (
    Boolean,
    Date,
    DateFormat,
    Email,
    ExcelAlchemy,
    ExcelMeta,
    FieldMeta,
    ImporterConfig,
    Number,
    NumberRange,
    Option,
    OptionId,
    Radio,
    String,
)
from excelalchemy._primitives.identity import UrlStr
from excelalchemy.core.storage_protocol import ExcelStorage
from excelalchemy.core.table import WorksheetTable
from excelalchemy.util.file import remove_excel_prefix

ROOT = Path(__file__).resolve().parents[1]
FILES_DIR = ROOT / 'files'
FILES_DIR.mkdir(exist_ok=True)
SHEET_NAME = 'Sheet1'

TEMPLATE_PATH = FILES_DIR / 'portfolio-template-en.xlsx'
INPUT_PATH = FILES_DIR / 'portfolio-import-input-en.xlsx'
RESULT_PATH = FILES_DIR / 'portfolio-import-result-en.xlsx'


def _load_table(path: Path, *, skiprows: int, sheet_name: str) -> WorksheetTable:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = [
            [None if value is None else str(value) for value in row]
            for row in worksheet.iter_rows(
                min_row=skiprows + 1,
                max_row=worksheet.max_row,
                max_col=worksheet.max_column,
                values_only=True,
            )
        ]

        while rows and all(value is None for value in rows[-1]):
            rows.pop()

        return WorksheetTable(rows=rows)
    finally:
        workbook.close()


class LocalPortfolioStorage(ExcelStorage):
    """Minimal local storage used to render screenshot assets on disk."""

    def read_excel_table(self, input_excel_name: str, *, skiprows: int, sheet_name: str) -> WorksheetTable:
        return _load_table(FILES_DIR / input_excel_name, skiprows=skiprows, sheet_name=sheet_name)

    def upload_excel(self, output_name: str, content_with_prefix: str) -> UrlStr:
        content = base64.b64decode(remove_excel_prefix(content_with_prefix))
        output_path = FILES_DIR / output_name
        output_path.write_bytes(content)
        return UrlStr(str(output_path))


TEAM_OPTIONS = [
    Option(id=OptionId('eng'), name='Engineering'),
    Option(id=OptionId('ops'), name='Operations'),
    Option(id=OptionId('sales'), name='Sales'),
]


class TemplateScreenshotImporter(BaseModel):
    full_name: String = FieldMeta(label='Full name', order=1, hint='Use the employee full name')
    age: Annotated[Number, Field(ge=18, le=65), ExcelMeta(label='Age', order=2, unit='years')]
    work_email: Email = FieldMeta(label='Work email', order=3, hint='Use a company email address')
    start_date: Date = FieldMeta(label='Start date', order=4, date_format=DateFormat.DAY)
    is_active: Boolean = FieldMeta(label='Status', order=5, hint='Yes for active employees, No otherwise')
    team: Radio = FieldMeta(label='Team', order=6, options=TEAM_OPTIONS)
    salary_band: NumberRange = FieldMeta(label='Salary band', order=7, unit='USD')


async def _creator(data: dict[str, object], context: None) -> dict[str, object]:
    return data


def _build_template_workbook() -> None:
    alchemy = ExcelAlchemy(
        ImporterConfig(
            TemplateScreenshotImporter,
            creator=_creator,
            storage=LocalPortfolioStorage(),
            locale='en',
        )
    )
    artifact = alchemy.download_template_artifact(
        sample_data=[
            {
                'full_name': 'Avery Stone',
                'age': 29,
                'work_email': 'avery.stone@example.com',
                'start_date': '2024-02-12',
                'is_active': True,
                'team': 'eng',
                'salary_band': {'start': 90000, 'end': 120000},
            }
        ],
        filename=TEMPLATE_PATH.name,
    )
    TEMPLATE_PATH.write_bytes(artifact.as_bytes())


def _build_invalid_input_workbook() -> None:
    shutil.copyfile(TEMPLATE_PATH, INPUT_PATH)
    workbook = load_workbook(INPUT_PATH)
    worksheet = workbook[SHEET_NAME]

    worksheet['A4'] = 'Taylor'
    worksheet['B4'] = '17'
    worksheet['C4'] = 'not-an-email'
    worksheet['D4'] = '2024-13-40'
    worksheet['E4'] = 'Maybe'
    worksheet['F4'] = 'Finance'
    worksheet['G4'] = '150000'
    worksheet['H4'] = '120000'

    workbook.save(INPUT_PATH)
    workbook.close()


def _build_result_workbook() -> None:
    alchemy = ExcelAlchemy(
        ImporterConfig(
            TemplateScreenshotImporter,
            creator=_creator,
            storage=LocalPortfolioStorage(),
            locale='en',
        )
    )
    asyncio.run(alchemy.import_data(INPUT_PATH.name, RESULT_PATH.name))


def main() -> None:
    _build_template_workbook()
    _build_invalid_input_workbook()
    _build_result_workbook()

    print(f'Generated template: {TEMPLATE_PATH}')
    print(f'Generated input: {INPUT_PATH}')
    print(f'Generated result: {RESULT_PATH}')


if __name__ == '__main__':
    main()
